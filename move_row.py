import openpyxl

def move_row_to_end(file_name, row_num):
    # Load the workbook and the active sheet
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active

    # Check if the row number is valid
    if row_num < 7:
        print("Cannot move rows before 7th row.")
        return

    # Extract the row data you want to move
    row_data = []
    for cell in sheet[row_num]:
        row_data.append(cell.value)

    # Delete the original row
    sheet.delete_rows(row_num)

    # Adjust the num column for all rows above the moved row starting from row 7
    for row in sheet.iter_rows(min_row=7, max_row=sheet.max_row, min_col=1, max_col=1):
      if row[0].row >= row_num:
        if row[0].value is not None:
            row[0].value -= 1
        


    # Set the num column value for the moved row to be last
    row_data[0] = sheet.max_row -6 # Subtract 6 because the first 6 rows aren't part of the numbering

    # Append the row data to the end of the sheet
    sheet.append(row_data)

    # Save the modified workbook
    workbook.save(file_name)

# Example usage:
file_name = '/home/genadi/Desktop/new(copy).xlsx'
row_to_move = 10  # Let's say you want to move the 5th row
move_row_to_end(file_name, row_to_move)
